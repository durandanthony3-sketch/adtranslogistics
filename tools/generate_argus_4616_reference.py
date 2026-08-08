#!/usr/bin/env python3
"""Génère le référentiel de traçabilité Argus 4616 du simulateur ADTL.

La source principale est le classeur d'extraction contrôlé du magazine. Les
prix marqués VALIDE_* alimentent les fourchettes de repli. Une ligne
A_VERIFIER ne peut être promue que lorsqu'une cote douanière observée la
recoupe exactement ; elle reste alors signalée comme « corroborée ».
"""

from __future__ import annotations

import argparse
import itertools
import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path(
    r"C:\Users\HP\Documents\New project 5\outputs\argus4616_20260711"
    r"\Extraction_ARGUS_4616_vehicules.xlsx"
)


def clean(value: object) -> str:
    text = unicodedata.normalize("NFD", str(value or "").upper())
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def norm(value: object) -> str:
    return clean(value).replace(" ", "")


def extract_json_assignment(text: str, pattern: str) -> dict:
    match = re.search(pattern, text, re.S)
    if not match:
        raise ValueError(f"Affectation JSON introuvable : {pattern}")
    return json.loads(match.group(1))


# Rapprochements issus des notes douanières et des séries 2023-2026.
# La relation sert uniquement à expliquer la source Argus ; elle n'écrase pas
# une cote exacte déjà présente dans le référentiel douanier observé.
MANUAL_LINKS: dict[tuple[str, str], tuple[str, str, str, str]] = {}


def add_link(
    source_brand: str,
    source_model: str,
    target_brand: str,
    target_model: str,
    relation: str,
    note: str,
) -> None:
    MANUAL_LINKS[(norm(source_brand), norm(source_model))] = (
        target_brand,
        target_model,
        relation,
        note,
    )


for source in ("RAV", "RAV4", "RAV 4", "TYPE RAV4"):
    add_link("TOYOTA", source, "TOYOTA", "RAV4", "alias", "Libellé rapproché de la gamme RAV4.")
add_link("TOYOTA", "COROLLA VERSO", "TOYOTA", "COROLLA VERSO", "alias", "Gamme Corolla Verso.")
for source in ("MATRIX", "MATRIX BASE"):
    add_link("TOYOTA", source, "TOYOTA", "COROLLA VERSO", "assimilation", "Matrix assimilée à Corolla Verso.")
for source in ("VIBE", "VIBE GT"):
    add_link("PONTIAC", source, "TOYOTA", "COROLLA VERSO", "assimilation", "Pontiac Vibe assimilée à Corolla Verso.")
for source in ("HIGHLANDER", "HIGHLANDER M C", "RUNNER", "4RUNNER"):
    add_link("TOYOTA", source, "TOYOTA", "LAND CRUISER", "assimilation", "Assimilation douanière à Toyota Land Cruiser.")
for source in ("CAMRY", "CAMRY M C"):
    add_link("TOYOTA", source, "TOYOTA", "AVENSIS", "assimilation", "Camry assimilée à Avensis (grande valeur selon les notes douanières).")
for source in ("AVALON", "SIENNA", "VENZA"):
    add_link("TOYOTA", source, "TOYOTA", "RAV4", "assimilation", f"{source.title()} assimilée à RAV4.")
add_link("PONTIAC", "TORRENT", "TOYOTA", "RAV4", "assimilation", "Pontiac Torrent assimilée à RAV4.")
for source in ("RX 330", "RX330", "RX350", "RX300"):
    add_link("LEXUS", source, "LEXUS", "RX", "alias", "Version rapprochée de la gamme Lexus RX.")
for source in ("CR V", "CR-V"):
    add_link("HONDA", source, "HONDA", "CR V", "alias", "Libellé rapproché de la gamme Honda CR-V.")
for source in ("MDX", "RDX"):
    add_link("ACURA", source, "HONDA", "CR V", "assimilation", f"Acura {source} assimilée à Honda CR-V.")
add_link("HONDA", "PILOT", "HONDA", "CR V", "assimilation", "Honda Pilot assimilé à Honda CR-V.")
add_link("ACURA", "TL", "HONDA", "ACCORD", "assimilation", "Acura TL assimilée à Honda Accord.")
for source in ("QASHQAI", "ROGUE"):
    add_link("NISSAN", source, "NISSAN", "X TRAIL", "assimilation", f"Nissan {source.title()} assimilé à X-Trail.")
add_link("NISSAN", "X TRAIL", "NISSAN", "X TRAIL", "alias", "Gamme Nissan X-Trail.")
add_link("MERCEDES BENZ", "ML350", "MERCEDES-BENZ", "CLASSE ML", "alias", "Version rapprochée de la Classe ML.")
add_link("MERCEDES BENZ", "GLK 350", "MERCEDES-BENZ", "CLASSE ML", "assimilation", "GLK 350 assimilé à ML 350.")
for source in ("C180", "C200", "C300"):
    add_link("MERCEDES BENZ", source, "MERCEDES-BENZ", "CLASSE C", "alias", "Version rapprochée de la Classe C.")
add_link("MERCEDES BENZ", "GLC300", "MERCEDES-BENZ", "GLC", "alias", "Version rapprochée de la gamme GLC.")
add_link("HONDA", "RIDGELINE", "TOYOTA", "HILUX", "assimilation", "Honda Ridgeline assimilé à Toyota Hilux.")
add_link("VOLKSWAGEN", "JETTA", "VOLKSWAGEN", "PASSAT", "assimilation", "Volkswagen Jetta assimilée à Passat.")
add_link("PORSCHE", "CAYENNE", "VOLKSWAGEN", "TOUAREG", "assimilation", "Porsche Cayenne assimilé à Volkswagen Touareg.")
add_link("INFINITI", "QX4", "NISSAN", "PATHFINDER", "assimilation", "Infiniti QX4 assimilé à Nissan Pathfinder.")
for source in ("J5", "J9"):
    add_link("PEUGEOT", source, "PEUGEOT", "BOXER", "assimilation", f"Peugeot {source} assimilé à Boxer.")
for number in ("3", "5", "6"):
    add_link("MAZDA", number, "MAZDA", f"MAZDA {number}", "alias", f"Libellé rapproché de Mazda {number}.")
    add_link("MAZDA", f"MAZDA {number}", "MAZDA", f"MAZDA {number}", "alias", f"Gamme Mazda {number}.")
add_link("LAND ROVER", "LANDROVER EVOQUE", "LAND ROVER", "RANGE ROVER EVOQUE", "alias", "Libellé rapproché de Range Rover Evoque.")


BRAND_ALIASES = {norm("MERCEDES BENZ"): "MERCEDES-BENZ"}


def resolve_link(brand: str, model: str) -> tuple[str, str, str, str]:
    manual = MANUAL_LINKS.get((norm(brand), norm(model)))
    if manual:
        return manual
    return BRAND_ALIASES.get(norm(brand), brand), model, "direct", "Correspondance directe de gamme."


def model_matches(source: str, target: str) -> bool:
    source_clean = clean(source)
    target_clean = clean(target)
    source_norm = norm(source_clean)
    target_norm = norm(target_clean)
    if not source_norm or not target_norm:
        return False
    if source_norm == target_norm:
        return True
    # L'OCR ajoute souvent une date de génération ou un code après le modèle.
    if source_clean.startswith(target_clean + " ") or source_norm.startswith(target_norm):
        # Éviter que CLASSE CLK/CLS/CLC soit capturé par la famille CLASSE C.
        if target_clean == "CLASSE C" and not (
            source_clean == "CLASSE C" or source_clean.startswith("CLASSE C ")
        ):
            return False
        # RX accepte les variantes RX300/RX350/RX400, mais pas RC/UX/NX.
        if target_clean == "RX" and not (
            source_clean == "RX"
            or source_clean.startswith("RX ")
            or re.match(r"^RX\d", source_clean)
        ):
            return False
        # Protéger les modèles très courts contre des rapprochements accidentels.
        if len(target_norm) <= 2 and source_norm != target_norm:
            return False
        return True
    return False


def line_payload(row: dict) -> dict:
    return {
        "value": int(round(row["value"])),
        "id": row["id"],
        "page": row["page_printed"],
        "pdfPage": row["page_pdf"],
        "version": row["version"],
        "status": row["status"],
    }


def find_value_proof(rows: list[dict], target: float) -> tuple[str | None, list[dict]]:
    by_value: dict[int, list[dict]] = defaultdict(list)
    for row in rows:
        by_value[int(round(row["value"]))].append(row)
    values = sorted(by_value)
    if not values:
        return None, []
    for value in values:
        if abs(value - target) <= 1:
            return "variant", [by_value[value][0]]
    low, high = values[0], values[-1]
    if abs(round((low + high) / 2) - target) <= 1:
        return "midrange", [by_value[low][0], by_value[high][0]]
    for small, large in itertools.combinations_with_replacement(values, 2):
        if abs(round((small + large) / 2) - target) <= 1:
            lines = [by_value[small][0]]
            if large != small:
                lines.append(by_value[large][0])
            return "pair", lines
    return None, []


def unique_rows(rows: list[dict]) -> list[dict]:
    seen: set[str] = set()
    result = []
    for row in rows:
        if row["id"] in seen:
            continue
        seen.add(row["id"])
        result.append(row)
    return result


def build_family(rows: list[dict]) -> dict[str, list]:
    years: dict[int, list[dict]] = defaultdict(list)
    for row in unique_rows(rows):
        if str(row["status"]).startswith("VALIDE"):
            years[row["year"]].append(row)
    result: dict[str, list] = {}
    for year, year_rows in sorted(years.items()):
        ordered = sorted(year_rows, key=lambda item: (item["value"], item["id"]))
        low, high = ordered[0], ordered[-1]
        result[str(year)] = [
            int(round(low["value"])),
            int(round(high["value"])),
            int(round((low["value"] + high["value"]) / 2)),
            ", ".join(str(value) for value in sorted({row["page_printed"] for row in ordered})),
            ", ".join(str(value) for value in sorted({row["page_pdf"] for row in ordered})),
            len(ordered),
            low["id"],
            high["id"],
            low["version"],
            high["version"],
        ]
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--simulateur", type=Path, default=Path("simulateur.html"))
    parser.add_argument("--referentiel", type=Path, default=Path("referentiel_douane.js"))
    parser.add_argument("--output", type=Path, default=Path("argus_4616_reference.js"))
    args = parser.parse_args()

    simulator_text = args.simulateur.read_text(encoding="utf-8")
    generic_argus = extract_json_assignment(
        simulator_text,
        r"<script>var ARGUS=(\{.*?\});</script>",
    )
    ref_text = args.referentiel.read_text(encoding="utf-8")
    customs_ref = extract_json_assignment(ref_text, r"var REF_DOUANE\s*=\s*(\{.*\});")

    selectable: set[tuple[str, str]] = set()
    for brand, models in generic_argus.items():
        for model in models:
            selectable.add((brand, model["modele"]))
    for brand, models in customs_ref.items():
        for model in models.values():
            selectable.add((brand, model["lbl"]))
    selectable.update((brand, model) for brand, model in MANUAL_LINKS)

    workbook = load_workbook(args.workbook, read_only=True, data_only=True)
    sheet = workbook["Extraction Argus"]
    iterator = sheet.iter_rows(values_only=True)
    headers = next(iterator)
    index = {value: idx for idx, value in enumerate(headers)}
    rows_by_brand: dict[str, list[dict]] = defaultdict(list)
    status_counts: dict[str, int] = defaultdict(int)
    for values in iterator:
        price = values[index["COTE_COURS_MOYEN_EUR"]]
        status = str(values[index["STATUT_PRIX"]] or "")
        status_counts[status] += 1
        if not isinstance(price, (int, float)) or price <= 0:
            continue
        brand = str(values[index["MARQUE"]] or "").strip()
        row = {
            "year": int(values[index["ANNEE_ARGUS"]]),
            "brand": brand,
            "canonical": str(values[index["MODELE_GAMME_CANONIQUE"]] or "").strip(),
            "raw": str(values[index["MODELE_GAMME_RAW"]] or "").strip(),
            "version": str(values[index["VERSION_FINITION_RAW"]] or "").strip(),
            "value": float(price),
            "status": status,
            "id": str(values[index["ID_VARIANTE"]] or ""),
            "page_printed": int(values[index["SOURCE_PAGE_IMPRIMEE"]]),
            "page_pdf": int(values[index["SOURCE_PAGE_PDF"]]),
        }
        rows_by_brand[norm(brand)].append(row)

    target_to_sources: dict[tuple[str, str], list[tuple[str, str]]] = defaultdict(list)
    links_output: dict[str, list[str]] = {}
    for source_brand, source_model in sorted(selectable):
        target_brand, target_model, relation, note = resolve_link(source_brand, source_model)
        target_key = (target_brand, target_model)
        target_to_sources[target_key].append((source_brand, source_model))
        links_output[f"{norm(source_brand)}|{norm(source_model)}"] = [
            target_brand,
            target_model,
            relation,
            note,
        ]

    family_rows: dict[tuple[str, str], list[dict]] = {}
    families_output: dict[str, dict] = {}
    for target_brand, target_model in sorted(target_to_sources):
        candidates = [
            row
            for row in rows_by_brand.get(norm(target_brand), [])
            if model_matches(row["canonical"], target_model)
            or model_matches(row["raw"], target_model)
        ]
        candidates = unique_rows(candidates)
        if not candidates:
            continue
        family_rows[(norm(target_brand), norm(target_model))] = candidates
        years = build_family(candidates)
        if years:
            families_output[f"{norm(target_brand)}|{norm(target_model)}"] = {
                "brand": target_brand,
                "model": target_model,
                "years": years,
            }

    trace_output: dict[str, dict] = {}
    trace_stats: dict[str, int] = defaultdict(int)
    for source_brand, models in customs_ref.items():
        for model in models.values():
            source_model = model["lbl"]
            target_brand, target_model, relation, note = resolve_link(source_brand, source_model)
            candidates = family_rows.get((norm(target_brand), norm(target_model)), [])
            candidate_years = sorted({row["year"] for row in candidates})
            for year_text, record in model["y"].items():
                if record[0] != "c" or not candidate_years:
                    continue
                year = int(year_text)
                retained = float(record[1])
                base_year = year
                coefficient = 1.0
                if year not in candidate_years and year < candidate_years[0]:
                    base_year = candidate_years[0]
                    gap = base_year - year
                    coefficient = 0.8 if gap == 1 else 0.6 if gap == 2 else 0.4
                year_rows = [row for row in candidates if row["year"] == base_year]
                target_value = retained / coefficient
                trusted_rows = [row for row in year_rows if str(row["status"]).startswith("VALIDE")]
                method, proof_rows = find_value_proof(trusted_rows, target_value)
                status = "verified"
                if not method:
                    method, proof_rows = find_value_proof(year_rows, target_value)
                    status = "corroborated" if method else "reference_only"
                if method and coefficient != 1:
                    method = "depreciated_" + method
                if not proof_rows and trusted_rows:
                    ordered = sorted(trusted_rows, key=lambda item: (item["value"], item["id"]))
                    proof_rows = [ordered[0]]
                    if ordered[-1]["id"] != ordered[0]["id"]:
                        proof_rows.append(ordered[-1])
                if not proof_rows:
                    continue
                trace_stats[status] += 1
                pages = sorted({row["page_printed"] for row in proof_rows})
                pdf_pages = sorted({row["page_pdf"] for row in proof_rows})
                values = sorted(row["value"] for row in proof_rows)
                trace_output[f"{norm(source_brand)}|{norm(source_model)}|{year}"] = {
                    "status": status,
                    "relation": relation,
                    "sourceBrand": target_brand,
                    "sourceModel": target_model,
                    "sourceYear": base_year,
                    "retained": int(round(retained)),
                    "baseValue": int(round(target_value)),
                    "coefficient": coefficient,
                    "method": method or "range",
                    "low": int(round(values[0])),
                    "high": int(round(values[-1])),
                    "pages": pages,
                    "pdfPages": pdf_pages,
                    "lines": [line_payload(row) for row in proof_rows],
                    "note": note,
                }

    meta = {
        "edition": "4616",
        "date": "29 juin 2022",
        "title": "L'argus n° 4616 du 29 juin 2022",
        "trustedRows": status_counts.get("VALIDE_RAPID", 0)
        + status_counts.get("VALIDE_CONSENSUS", 0),
        "toReviewRows": status_counts.get("A_VERIFIER", 0),
        "years": [2008, 2021],
        "missingPrintedPages": [122, 123],
        "rule": "Cote de base rapprochée de l'Argus ; dépréciation 80 % / 60 % / 40 % sous le premier millésime.",
        "generatedFrom": args.workbook.name,
        "traceStats": dict(trace_stats),
    }

    banner = (
        "// Généré par tools/generate_argus_4616_reference.py — ne pas modifier à la main.\n"
        "// Source : L'argus n° 4616 du 29 juin 2022 ; seules les lignes traçables sont exposées.\n"
    )
    output = banner
    output += "window.ARGUS_4616_META=" + json.dumps(meta, ensure_ascii=False, separators=(",", ":")) + ";\n"
    output += "window.ARGUS_4616_FAMILIES=" + json.dumps(families_output, ensure_ascii=False, separators=(",", ":")) + ";\n"
    output += "window.ARGUS_4616_LINKS=" + json.dumps(links_output, ensure_ascii=False, separators=(",", ":")) + ";\n"
    output += "window.ARGUS_4616_TRACE=" + json.dumps(trace_output, ensure_ascii=False, separators=(",", ":")) + ";\n"
    args.output.write_text(output, encoding="utf-8", newline="\n")
    print(
        json.dumps(
            {
                "output": str(args.output.resolve()),
                "families": len(families_output),
                "links": len(links_output),
                "traces": len(trace_output),
                "traceStats": dict(trace_stats),
                "bytes": args.output.stat().st_size,
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
