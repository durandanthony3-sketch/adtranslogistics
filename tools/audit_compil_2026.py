#!/usr/bin/env python3
"""Audit global du simulateur ADTL contre les liquidations payées de 2026.

Le script est volontairement en lecture seule pour le classeur source. Il
normalise les libellés de véhicules, rapproche les modèles proposés par le
simulateur et vérifie la formule sur les catégories fiscales A à D.
"""

from __future__ import annotations

import argparse
import json
import math
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path(
    r"C:\Users\HP\Documents\Claude\Projects\AD TRANS LOGISTICS\COMPIL DEDOUANEMENT.xlsx"
)
RATES = {"A": 0.20, "B": 0.20, "C": 0.10, "D": 0.05}
FIXED_TAXES = 98_600


def clean(value: object) -> str:
    text = unicodedata.normalize("NFD", str(value or "").upper())
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def norm(value: object) -> str:
    return clean(value).replace(" ", "")


def js_round(value: float) -> int:
    return math.floor(value + 0.5)


def liquidate(caf: int, category: str) -> int:
    assiette = 0.70 * caf
    rate = RATES[category]
    duty = js_round(assiette * rate)
    community = js_round(assiette * 0.03)
    tax_base = assiette * (1 + rate + 0.03)
    vat = js_round(tax_base * 0.18)
    excise = js_round(tax_base * 0.10) if category == "A" else 0
    return duty + community + vat + excise + FIXED_TAXES


def extract_json_assignment(text: str, pattern: str) -> dict:
    match = re.search(pattern, text, re.S)
    if not match:
        raise ValueError(f"Affectation JSON introuvable : {pattern}")
    return json.loads(match.group(1))


def load_catalog(root: Path) -> tuple[dict[str, list[dict]], dict]:
    simulator = (root / "simulateur.html").read_text(encoding="utf-8")
    generic = extract_json_assignment(simulator, r"<script>var ARGUS=(\{.*?\});</script>")
    customs = extract_json_assignment(
        (root / "referentiel_douane.js").read_text(encoding="utf-8"),
        r"var REF_DOUANE\s*=\s*(\{.*\});",
    )
    catalog: dict[str, list[dict]] = defaultdict(list)
    for brand, models in generic.items():
        for model in models:
            catalog[brand].append({"key": model["modele"], "label": model["modele"]})
    for brand, models in customs.items():
        for key, model in models.items():
            catalog[brand].append({"key": key, "label": model["lbl"]})
    for brand in catalog:
        unique = {}
        for model in catalog[brand]:
            unique[norm(model["label"])] = model
            unique[norm(model["key"])] = model
        catalog[brand] = list({item["key"]: item for item in unique.values()}.values())
    return dict(catalog), customs


BRAND_ALIASES = {
    "MERCEDES": "MERCEDES BENZ",
    "MERCEDESBENZ": "MERCEDES BENZ",
    "VOLKSWAGEN": "VOLKSWAGEN",
    "VW": "VOLKSWAGEN",
    "LANDROVER": "LAND ROVER",
    "RANGEROVER": "LAND ROVER",
    "GM": "GMC",
}

MODEL_ALIASES = {
    ("TOYOTA", "RAV"): "RAV4",
    ("TOYOTA", "RAV4"): "RAV4",
    ("TOYOTA", "4RUNNER"): "4RUNNER",
    ("HONDA", "CRV"): "CR-V",
    ("NISSAN", "XTRAIL"): "X TRAIL",
    ("MERCEDES BENZ", "ML350"): "ML350",
    ("MERCEDES BENZ", "GLK350"): "GLK 350",
    ("LAND ROVER", "RANGEROVEREVOQUE"): "LANDROVER EVOQUE",
}

NOISE = {
    "MC",
    "M",
    "C",
    "MODELE",
    "MODEL",
    "TYPE",
    "ANNEE",
    "ESSENCE",
    "DIESEL",
    "AUTOMATIQUE",
    "MANUELLE",
    "USAGE",
    "USAGES",
}


def detect_brand(raw: str, catalog: dict[str, list[dict]]) -> str | None:
    compact = norm(raw)
    candidates = []
    for brand in catalog:
        brand_norm = norm(brand)
        if brand_norm and brand_norm in compact:
            candidates.append((len(brand_norm), brand))
    for alias, brand in BRAND_ALIASES.items():
        if alias in compact and brand in catalog:
            candidates.append((len(alias), brand))
    return max(candidates)[1] if candidates else None


def detect_year(raw: str) -> int | None:
    years = [int(value) for value in re.findall(r"(?<!\d)(19[6-9]\d|20[0-2]\d)(?!\d)", raw)]
    years = [year for year in years if 1960 <= year <= 2026]
    return years[-1] if years else None


def model_segment(raw: str, brand: str, year: int) -> str:
    text = clean(raw)
    text = re.sub(rf"\b{year}\b", " ", text)
    for token in sorted({clean(brand), "MERCEDES BENZ", "LAND ROVER", "RANGE ROVER"}, key=len, reverse=True):
        if token:
            text = re.sub(rf"\b{re.escape(token)}\b", " ", text)
    words = [word for word in text.split() if word not in NOISE]
    return " ".join(words)


def detect_model(raw: str, brand: str, year: int, catalog: dict[str, list[dict]]) -> dict | None:
    segment = model_segment(raw, brand, year)
    compact = norm(segment)
    alias = MODEL_ALIASES.get((brand, compact))
    if alias:
        alias_norm = norm(alias)
        for item in catalog[brand]:
            if alias_norm in {norm(item["key"]), norm(item["label"])}:
                return item
    candidates = []
    for item in catalog[brand]:
        for value in {item["key"], item["label"]}:
            candidate = norm(value)
            if len(candidate) >= 2 and candidate in compact:
                candidates.append((len(candidate), item))
    return max(candidates, key=lambda value: value[0])[1] if candidates else None


def zone_from_origin(origin: object) -> str:
    value = clean(origin)
    if any(token in value for token in ("ETATS UNIS", "USA", "CANADA")):
        return "usa"
    if any(token in value for token in ("JAPON", "CHINE", "COREE", "THAILANDE", "INDE")):
        return "asie"
    if any(token in value for token in ("DUBAI", "EMIRATS", "ARABIE")):
        return "dubai"
    if any(token in value for token in ("BENIN", "TOGO", "NIGERIA", "GHANA", "AFRIQUE")):
        return "afrique"
    return "europe"


def current_reference(customs: dict, brand: str, model_key: str, year: int) -> list | None:
    model = customs.get(brand, {}).get(model_key)
    if not model:
        return None
    return model.get("y", {}).get(str(year))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--root", type=Path, default=Path(__file__).resolve().parents[1])
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()

    catalog, customs = load_catalog(args.root)
    workbook = load_workbook(args.workbook, read_only=True, data_only=True)
    sheet = workbook["2026"]
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows)
    index = {str(value): position for position, value in enumerate(headers)}

    paid = 0
    numeric = 0
    vehicle_rows = 0
    parsed_rows = 0
    formula_exact = 0
    formula_close = 0
    categories = Counter()
    unmatched = Counter()
    unmatched_examples: dict[str, list[str]] = defaultdict(list)
    formula_mismatches = []
    profiles: dict[str, dict[tuple[int, int, str], dict]] = defaultdict(dict)
    reference_coverage = Counter()

    for excel_row, values in enumerate(rows, start=2):
        if values[index["Statut"]] != "ST_PAID":
            continue
        paid += 1
        caf = values[index["Val. CAF"]]
        taxes = values[index["Tot. Liquidé"]]
        if not isinstance(caf, (int, float)) or not isinstance(taxes, (int, float)):
            continue
        caf = int(round(caf))
        taxes = int(round(taxes))
        if caf <= 0 or taxes <= 0:
            continue
        numeric += 1
        raw = str(values[index["Marque"]] or "").strip()
        brand = detect_brand(raw, catalog)
        year = detect_year(raw)
        if not brand or not year:
            continue
        vehicle_rows += 1
        model = detect_model(raw, brand, year, catalog)
        if not model:
            key = f"{brand}|{model_segment(raw, brand, year)}|{year}"
            unmatched[key] += 1
            if len(unmatched_examples[key]) < 3:
                unmatched_examples[key].append(raw)
            continue
        parsed_rows += 1
        differences = {category: abs(liquidate(caf, category) - taxes) for category in RATES}
        category = min(differences, key=differences.get)
        delta = differences[category]
        if delta <= 3:
            formula_exact += 1
        elif len(formula_mismatches) < 100:
            formula_mismatches.append(
                {
                    "row": excel_row,
                    "vehicle": raw,
                    "brand": brand,
                    "model": model["label"],
                    "year": year,
                    "caf": caf,
                    "observedTaxes": taxes,
                    "closestCategory": category,
                    "calculatedTaxes": liquidate(caf, category),
                    "delta": delta,
                    "allCategories": {
                        candidate: liquidate(caf, candidate) for candidate in RATES
                    },
                }
            )
        if delta <= 100:
            formula_close += 1
        categories[category] += 1
        key = f"{brand}|{model['key']}|{year}"
        signature = (caf, taxes, category)
        record = profiles[key].setdefault(
            signature,
            {
                "caf": caf,
                "taxes": taxes,
                "category": category,
                "count": 0,
                "maxDelta": 0,
                "origins": Counter(),
                "rows": [],
                "label": model["label"],
            },
        )
        record["count"] += 1
        record["maxDelta"] = max(record["maxDelta"], delta)
        record["origins"][zone_from_origin(values[index["Pays d’origine"]])] += 1
        if len(record["rows"]) < 5:
            record["rows"].append(excel_row)

        ref = current_reference(customs, brand, model["key"], year)
        reference_coverage["present" if ref else "missing"] += 1

    profile_output = {}
    for key, records in sorted(profiles.items()):
        values = []
        for record in records.values():
            item = dict(record)
            item["origins"] = dict(record["origins"])
            values.append(item)
        values.sort(key=lambda item: (item["taxes"], item["caf"]), reverse=True)
        profile_output[key] = values

    report = {
        "workbook": str(args.workbook),
        "sheet": "2026",
        "rows": sheet.max_row - 1,
        "paidRows": paid,
        "numericPaidRows": numeric,
        "vehicleRowsDetected": vehicle_rows,
        "vehicleRowsMatched": parsed_rows,
        "vehicleMatchRate": round(parsed_rows / vehicle_rows, 4) if vehicle_rows else 0,
        "formulaExactWithin3": formula_exact,
        "formulaExactRate": round(formula_exact / parsed_rows, 4) if parsed_rows else 0,
        "formulaWithin100": formula_close,
        "categories": dict(categories),
        "matchedVehicleYearFamilies": len(profile_output),
        "referenceCoverageRows": dict(reference_coverage),
        "formulaMismatches": formula_mismatches,
        "unmatchedFamilies": [
            {"key": key, "count": count, "examples": unmatched_examples[key]}
            for key, count in unmatched.most_common()
        ],
        "profiles": profile_output,
    }
    payload = json.dumps(report, ensure_ascii=False, indent=2)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(payload, encoding="utf-8")
    print(json.dumps({key: value for key, value in report.items() if key != "profiles"}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
